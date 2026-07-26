from pathlib import Path
from typing import List
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from src.models.result import Result
from src.utils.logger import logger
from datetime import datetime


class ExcelExporter:
    def __init__(self, output_dir: str = "results"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def _generate_date_filename(self, filename: str) -> str:
        """根据日期生成文件名，如 evaluation_results_20260726.xlsx"""
        today = datetime.now()
        date_str = today.strftime("%Y%m%d")
        
        base_name = Path(filename).stem
        ext = Path(filename).suffix or ".xlsx"
        
        return f"{base_name}_{date_str}{ext}"
    
    def export(self, results: List[Result], filename: str = "evaluation_results.xlsx") -> str:
        if not results:
            logger.warning("没有结果可导出")
            return ""
        
        date_filename = self._generate_date_filename(filename)
        filepath = self.output_dir / date_filename
        
        file_exists = filepath.exists()
        
        if file_exists:
            logger.info(f"文件 {date_filename} 已存在，将追加数据")
        else:
            logger.info(f"文件 {date_filename} 不存在，将创建新文件")
        
        data = []
        for result in results:
            image_note = ""
            if result.screenshot_path:
                if getattr(result, 'is_shared_image', False):
                    image_note = "分享图片"
                else:
                    image_note = "页面截图（可手动替换）"
            
            data.append({
                "问题ID": result.question_id,
                "平台": result.platform_name,
                "问题": result.question_text,
                "回答": result.answer,
                "状态": result.status,
                "匹配规则": ", ".join(result.matched_rules) if result.matched_rules else "",
                "错误信息": result.error_message or "",
                "时间": result.timestamp.strftime("%Y-%m-%d %H:%M:%S") if result.timestamp else "",
                "截图类型": image_note,
                "截图路径": result.screenshot_path or "",
                "分享链接": result.share_link or "",
                "分享链接失败原因": result.share_link_error or ""
            })
        
        columns = list(data[0].keys()) if data else []
        
        if file_exists:
            self._append_to_existing_file(filepath, data, columns, results)
        else:
            self._create_new_file(filepath, data, columns, results)
        
        self._export_share_links(results)
        
        logger.info(f"Excel报告已导出: {filepath}")
        return str(filepath)
    
    def _create_new_file(self, filepath: Path, data: List[dict], columns: list, results: List[Result]):
        """创建新的Excel文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "评测结果"
        
        self._write_header(ws, columns)
        self._write_data_rows(ws, data, results, start_row=2)
        self._format_columns(ws)
        
        wb.save(filepath)
        logger.info(f"新Excel文件已创建: {filepath}")
    
    def _append_to_existing_file(self, filepath: Path, data: List[dict], columns: list, results: List[Result]):
        """追加数据到已存在的Excel文件"""
        wb = load_workbook(filepath)
        ws = wb.active
        
        max_row = ws.max_row
        start_row = max_row + 1
        
        logger.info(f"从第 {start_row} 行开始追加数据")
        
        self._write_data_rows(ws, data, results, start_row=start_row)
        
        wb.save(filepath)
        logger.info(f"数据已追加到Excel文件: {filepath}")
    
    def _write_header(self, ws, columns: list):
        """写入表头并设置样式"""
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _write_data_rows(self, ws, data: List[dict], results: List[Result], start_row: int):
        """写入数据行和截图"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        columns = list(data[0].keys()) if data else []
        
        for row_offset, (row_data, result) in enumerate(zip(data, results), start=0):
            current_row = start_row + row_offset
            
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=row_data.get(col_name, ""))
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            if result.screenshot_path:
                screenshot_path = Path(result.screenshot_path)
                if not screenshot_path.is_absolute():
                    screenshot_path = Path.cwd() / result.screenshot_path
                
                if screenshot_path.exists():
                    try:
                        img = OpenpyxlImage(str(screenshot_path))
                        img.width = 200
                        img.height = 150
                        
                        img_cell = f'J{current_row}'
                        ws.add_image(img, img_cell)
                        
                        ws.row_dimensions[current_row].height = 120
                        logger.info(f"截图成功插入到单元格 {img_cell}")
                    except Exception as e:
                        logger.error(f"无法插入截图 {screenshot_path}: {str(e)}")
                else:
                    logger.warning(f"截图文件不存在: {screenshot_path}")
    
    def _format_columns(self, ws):
        """设置列宽（仅在创建新文件时调用）"""
        column_widths = {
            'A': 12, 'B': 15, 'C': 40, 'D': 60, 'E': 10, 'F': 20,
            'G': 30, 'H': 20, 'I': 20, 'J': 40, 'K': 50, 'L': 50
        }
        
        for col, width in column_widths.items():
            if col in ws.column_dimensions:
                ws.column_dimensions[col].width = width
    
    def _export_share_links(self, results: List[Result]):
        """将分享链接追加到文本文件，保持和图片名称一致性"""
        links_filepath = self.output_dir / "share_links.txt"
        
        try:
            file_exists = links_filepath.exists()
            
            with open(links_filepath, 'a', encoding='utf-8') as f:
                if not file_exists:
                    f.write("=" * 80 + "\n")
                    f.write("AI 问答评测工具 - 分享链接汇总\n")
                    f.write("=" * 80 + "\n\n")
                
                for result in results:
                    if result.share_link:
                        timestamp = result.timestamp.strftime("%Y%m%d_%H%M%S") if result.timestamp else datetime.now().strftime("%Y%m%d_%H%M%S")
                        screenshot_name = f"{result.platform_name}_{result.question_id}_{timestamp}"
                        
                        f.write(f"[{result.platform_name}] 问题 {result.question_id} | 文件名: {screenshot_name}\n")
                        f.write(f"  问题: {result.question_text}\n")
                        f.write(f"  链接: {result.share_link}\n")
                        f.write("-" * 80 + "\n")
            
            logger.info(f"分享链接已追加到: {links_filepath}")
        except Exception as e:
            logger.error(f"导出分享链接失败: {str(e)}")
    
    def export_summary(self, results: List[Result], filename: str = "summary.xlsx") -> str:
        if not results:
            logger.warning("没有结果可导出")
            return ""
        
        date_filename = self._generate_date_filename(filename)
        filepath = self.output_dir / date_filename
        
        file_exists = filepath.exists()
        
        summary_data = {
            "总结果数": len(results),
            "成功数": sum(1 for r in results if r.status == "success"),
            "失败数": sum(1 for r in results if r.status == "error"),
            "成功率": f"{sum(1 for r in results if r.status == 'success') / len(results) * 100:.2f}%"
        }
        
        platforms = set(r.platform_name for r in results)
        platform_stats = {}
        
        for platform in platforms:
            platform_results = [r for r in results if r.platform_name == platform]
            platform_stats[platform] = {
                "总数": len(platform_results),
                "成功": sum(1 for r in platform_results if r.status == "success"),
                "失败": sum(1 for r in platform_results if r.status == "error"),
                "成功率": f"{sum(1 for r in platform_results if r.status == 'success') / len(platform_results) * 100:.2f}%"
            }
        
        df_summary = pd.DataFrame([summary_data])
        df_platforms = pd.DataFrame(platform_stats).T
        
        if file_exists:
            wb = load_workbook(filepath)
            if '总览' in wb.sheetnames:
                ws = wb['总览']
                max_row = ws.max_row
                for _, row in df_summary.iterrows():
                    ws.append(row.tolist())
            if '平台统计' in wb.sheetnames:
                ws = wb['平台统计']
                max_row = ws.max_row
                for idx, row in df_platforms.iterrows():
                    ws.append([idx] + row.tolist())
            wb.save(filepath)
        else:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df_summary.to_excel(writer, sheet_name='总览', index=False)
                df_platforms.to_excel(writer, sheet_name='平台统计')
        
        logger.info(f"汇总报告已导出: {filepath}")
        return str(filepath)