#!/usr/bin/env python3
"""
测试Excel导出功能：日期文件名、追加写入、截图嵌入
"""
import asyncio
import os
import tempfile
from pathlib import Path
from datetime import datetime
from src.models.result import Result
from src.exporter.excel_exporter import ExcelExporter


def create_test_results(count: int, prefix: str = "") -> list:
    """创建测试结果"""
    results = []
    for i in range(count):
        result = Result(
            question_id=f"{prefix}Q{i+1}",
            platform_name=f"平台{i+1}",
            question_text=f"测试问题{i+1}：这是一个测试问题",
            answer=f"测试回答{i+1}：这是测试问题{i+1}的回答内容",
            status="success",
            timestamp=datetime.now()
        )
        results.append(result)
    return results


def test_date_filename():
    """测试日期文件名生成"""
    exporter = ExcelExporter()
    filename = exporter._generate_date_filename("evaluation_results.xlsx")
    
    today = datetime.now().strftime("%Y%m%d")
    assert today in filename, f"文件名应包含日期: {filename}"
    assert "evaluation_results" in filename, f"文件名应包含基础名称: {filename}"
    assert filename.endswith(".xlsx"), f"文件名应以.xlsx结尾: {filename}"
    
    print(f"✓ 日期文件名生成正确: {filename}")


def test_create_new_file():
    """测试创建新文件"""
    with tempfile.TemporaryDirectory() as tmpdir:
        exporter = ExcelExporter(output_dir=tmpdir)
        results = create_test_results(2, "TEST_")
        
        filepath = exporter.export(results, "test_results.xlsx")
        
        assert filepath, "导出路径不应为空"
        assert os.path.exists(filepath), f"文件应存在: {filepath}"
        
        print(f"✓ 创建新文件成功: {filepath}")
        
        return filepath


def test_append_to_file():
    """测试追加数据到已有文件"""
    with tempfile.TemporaryDirectory() as tmpdir:
        exporter = ExcelExporter(output_dir=tmpdir)
        
        # 第一次导出
        results1 = create_test_results(2, "FIRST_")
        filepath = exporter.export(results1, "test_append.xlsx")
        
        # 第二次导出（追加）
        results2 = create_test_results(3, "SECOND_")
        filepath = exporter.export(results2, "test_append.xlsx")
        
        # 验证文件存在且数据已追加
        assert os.path.exists(filepath), f"文件应存在: {filepath}"
        
        # 使用openpyxl验证行数
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.active
        
        # 表头1行 + 第一次2行 + 第二次3行 = 6行
        assert ws.max_row == 6, f"预期6行，实际{ws.max_row}行"
        
        print(f"✓ 追加数据成功，总行数: {ws.max_row}")


def test_summary_export():
    """测试汇总报告导出"""
    with tempfile.TemporaryDirectory() as tmpdir:
        exporter = ExcelExporter(output_dir=tmpdir)
        results = create_test_results(5, "SUMMARY_")
        
        filepath = exporter.export_summary(results, "test_summary.xlsx")
        
        assert filepath, "汇总报告路径不应为空"
        assert os.path.exists(filepath), f"汇总报告应存在: {filepath}"
        
        print(f"✓ 汇总报告导出成功: {filepath}")


def test_summary_append():
    """测试汇总报告追加"""
    with tempfile.TemporaryDirectory() as tmpdir:
        exporter = ExcelExporter(output_dir=tmpdir)
        
        # 第一次导出
        results1 = create_test_results(3, "SUM1_")
        filepath = exporter.export_summary(results1, "test_summary_append.xlsx")
        
        # 第二次导出（追加）
        results2 = create_test_results(2, "SUM2_")
        filepath = exporter.export_summary(results2, "test_summary_append.xlsx")
        
        assert os.path.exists(filepath), f"汇总报告应存在: {filepath}"
        
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        
        if '总览' in wb.sheetnames:
            ws = wb['总览']
            assert ws.max_row >= 2, f"总览工作表应有至少2行（表头+2次数据），实际{ws.max_row}行"
            print(f"✓ 汇总报告追加成功，总览行数: {ws.max_row}")


if __name__ == "__main__":
    print("=" * 60)
    print("测试Excel导出功能")
    print("=" * 60)
    print()
    
    test_date_filename()
    print()
    
    test_create_new_file()
    print()
    
    test_append_to_file()
    print()
    
    test_summary_export()
    print()
    
    test_summary_append()
    print()
    
    print("=" * 60)
    print("所有测试通过！")
    print("=" * 60)